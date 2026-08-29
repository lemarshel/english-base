#!/usr/bin/env python
"""Convert the COCA-style lemma frequency spreadsheet into the site's word data.

Reads an .xlsx with the COCA lemma columns (rank, lemma, PoS, perMil,
dispersion, range, normalized genre frequencies, and capitalization signals)
and writes:

    data/words.json      the vocabulary as plain JSON
    data/words-data.js   the same records as the two globals index.html loads

Translations are merged from data/ru and data/kk. Definition and example fields
are emitted empty — the slots exist so they can be filled in later without
touching the schema.

    python tools/build_words.py ~/Downloads/lemmas_60k.xlsx --limit 5000
"""
import argparse
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# COCA part-of-speech codes -> the nine sections the interface renders.
POS_MAP = {
    'n': 'pos_noun',
    'v': 'pos_verb',
    'j': 'pos_adj',
    'r': 'pos_adv',
    'e': 'pos_adv',      # existential "there"
    'm': 'pos_det',      # numerals behave as determiners in English
    'a': 'pos_det',      # articles
    'd': 'pos_det',
    'u': 'pos_interj',
    'c': 'pos_conj',
    'i': 'pos_prep',
    'p': 'pos_pron',
}

# Section order, matching the table of contents in js/render-words.js.
POS_ORDER = ['pos_noun', 'pos_verb', 'pos_adj', 'pos_adv', 'pos_det',
             'pos_interj', 'pos_conj', 'pos_prep', 'pos_pron']

# Only normalized per-million columns are retained for genre comparisons. Raw
# genre counts vary with corpus size and are therefore unsuitable for comparing
# spoken, fiction, academic, news, web, and other registers.
COCA_NUMERIC_FIELDS = {
    'perMil': 'per_mil',
    '%caps': 'caps_pct',
    '%allC': 'all_caps_pct',
    'range': 'range',
    'disp': 'disp',
    'blogPM': 'blog_pm',
    'webPM': 'web_pm',
    'TVMPM': 'tvm_pm',
    'spokPM': 'spok_pm',
    'ficPM': 'fic_pm',
    'magPM': 'mag_pm',
    'newsPM': 'news_pm',
    'acadPM': 'acad_pm',
}

# CEFR is not in the source, so it is approximated from frequency rank. These
# bands are a rough proxy, not measured levels — replace them when real CEFR
# data is available.
CEFR_BANDS = [(1000, 'a1'), (2500, 'a2'), (5000, 'b1'),
              (10000, 'b2'), (20000, 'c1')]


def cefr_for(rank):
    for limit, level in CEFR_BANDS:
        if rank <= limit:
            return level
    return 'c2'


def load_translations(kind):
    """Merge every data/<kind>/*.json into {pos: {word: translation}}.

    Translations live apart from the generated word data so the list can be
    rebuilt from the spreadsheet without losing them, and so each part of speech
    can be worked on in its own file.
    """
    merged = {}
    folder = ROOT / 'data' / kind
    if not folder.is_dir():
        return merged
    for path in sorted(folder.glob('*.json')):
        with path.open(encoding='utf-8') as f:
            data = json.load(f)
        for pos, entries in data.items():
            merged.setdefault(pos, {}).update(entries)
    return merged


def usable(lemma):
    """Skip slashed alternatives and date compounds — 'his/her', 'mid-1930s'."""
    if not lemma:
        return False
    text = str(lemma).strip()
    if not text or '/' in text or any(ch.isdigit() for ch in text):
        return False
    return bool(re.match(r"^[A-Za-z][A-Za-z'\- ]*$", text))


def number(value, default=0):
    """Return a JSON-friendly int/float for a possibly empty spreadsheet cell."""
    if value in (None, ''):
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value
    try:
        parsed = float(value)
        return int(parsed) if parsed.is_integer() else parsed
    except (TypeError, ValueError):
        return default


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('--limit', type=int, default=5000)
    ap.add_argument('--sheet', default=None)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    values = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else '' for value in next(values)]
    required = {'rank', 'lemma', 'PoS'} | set(COCA_NUMERIC_FIELDS)
    missing = sorted(required - set(headers))
    if missing:
        ap.error('spreadsheet is missing COCA columns: ' + ', '.join(missing))

    rows = []
    for values_row in values:
        source = dict(zip(headers, values_row))
        rank = source.get('rank')
        lemma = source.get('lemma')
        pos = source.get('PoS')
        if rank is None or not usable(lemma):
            continue
        rows.append({
            'rank': int(rank),
            'lemma': str(lemma).strip(),
            'coca_pos': str(pos or '').strip().lower(),
            **{
                target: number(source.get(column))
                for column, target in COCA_NUMERIC_FIELDS.items()
            },
        })

    rows.sort(key=lambda row: row['rank'])  # most frequent first
    rows = rows[:args.limit]

    ru = load_translations('ru')
    kk = load_translations('kk')

    words = []
    for i, source in enumerate(rows, start=1):
        rank = source['rank']
        lemma = source['lemma']
        coca_pos = source['coca_pos']
        section = POS_MAP.get(coca_pos, 'pos_noun')
        words.append({
            'id': i,
            'word': lemma,
            'ipa': '',
            'ru': ru.get(section, {}).get(lemma, ''),
            'kk': kk.get(section, {}).get(lemma, ''),
            'def': '',
            'pos': section,
            'group': 'g_' + section.replace('pos_', ''),
            'root': '',
            'affix': '',
            'cefr': cefr_for(rank),
            'ex': '',
            'ex_ipa': '',
            'ex_ru': '',
            'ex_kk': '',
            'ex_def': '',
            'rank': rank,
            'coca_pos': coca_pos,
            'per_mil': source['per_mil'],
            'disp': source['disp'],
            'range': source['range'],
            'spok_pm': source['spok_pm'],
            'tvm_pm': source['tvm_pm'],
            'fic_pm': source['fic_pm'],
            'acad_pm': source['acad_pm'],
            'news_pm': source['news_pm'],
            'mag_pm': source['mag_pm'],
            'web_pm': source['web_pm'],
            'blog_pm': source['blog_pm'],
            'caps_pct': source['caps_pct'],
            'all_caps_pct': source['all_caps_pct'],
        })

    # Group by first letter inside each part of speech.
    #
    # The size matters as much as the grouping: the row virtualizer estimates a
    # tbody's height from a single measured row, so a 2400-row tbody turns a
    # few pixels of error into thousands, and the page height thrashes while
    # scrolling. Small groups keep that error bounded — the same shape hsk-base
    # has with its phonetic groups. Headers are hidden by default anyway.
    words.sort(key=lambda w: (POS_ORDER.index(w['pos']), w['word'][0].lower(), w['rank']))

    groups = []
    seen_pos = set()
    for w in words:
        letter = w['word'][0].upper()
        gid = 'g_%s_%s' % (w['pos'].replace('pos_', ''), letter.lower())
        w['group'] = gid
        if gid in {g['id'] for g in groups}:
            continue
        first_of_section = w['pos'] not in seen_pos
        seen_pos.add(w['pos'])
        groups.append({
            'id': gid,
            'label': letter,          # rendered instead of a root
            'root': '',
            'root_key': '',
            'affix': '',
            'affix_key': '',
            'standalone': False,
            'pos_heading_before': w['pos'] if first_of_section else None,
        })

    for i, w in enumerate(words, start=1):
        w['id'] = i

    (ROOT / 'data').mkdir(exist_ok=True)

    # One record per line: valid JSON, but diffable and a third the size of
    # a fully indented dump.
    def compact(obj):
        return json.dumps(obj, ensure_ascii=False, separators=(',', ':'))

    with (ROOT / 'data' / 'words.json').open('w', encoding='utf-8') as f:
        f.write('{\n"groups": [\n')
        for i, g in enumerate(groups):
            f.write(compact(g) + (',' if i < len(groups) - 1 else '') + '\n')
        f.write('],\n"words": [\n')
        for i, w in enumerate(words):
            f.write(compact(w) + (',' if i < len(words) - 1 else '') + '\n')
        f.write(']\n}\n')

    header = (
        '/* ==========================================================================\n'
        '   English Base — word data\n'
        '   --------------------------------------------------------------------------\n'
        '   GENERATED by tools/build_words.py — do not hand-edit the arrays below;\n'
        '   edit data/words.json and regenerate, or extend the build script.\n'
        '\n'
        '   ru / kk translations are merged from data/ru and data/kk. Definition\n'
        '   and example fields remain available for later enrichment.\n'
        '   Field reference: docs/architecture.md\n'
        '   ========================================================================== */\n'
    )
    with (ROOT / 'data' / 'words-data.js').open('w', encoding='utf-8') as f:
        f.write(header)
        f.write('window.EN_GROUPS = ' + json.dumps(groups, ensure_ascii=False) + ';\n')
        f.write('window.EN_WORDS = [\n')
        for i, w in enumerate(words):
            comma = ',' if i < len(words) - 1 else ''
            f.write(json.dumps(w, ensure_ascii=False, separators=(',', ':')) + comma + '\n')
        f.write('];\n')

    counts = {}
    for w in words:
        counts[w['pos']] = counts.get(w['pos'], 0) + 1
    levels = {}
    for w in words:
        levels[w['cefr']] = levels.get(w['cefr'], 0) + 1

    print('words: %d   rank span: %d-%d'
          % (len(words), min(w['rank'] for w in words), max(w['rank'] for w in words)))
    print('sections:', {k: counts[k] for k in POS_ORDER if k in counts})
    print('cefr:', {k: levels[k] for k in ['a1', 'a2', 'b1', 'b2', 'c1', 'c2'] if k in levels})

    for lang in ('ru', 'kk'):
        done = sum(1 for w in words if w[lang])
        if done or lang == 'ru':
            missing = {}
            for w in words:
                if not w[lang]:
                    missing[w['pos']] = missing.get(w['pos'], 0) + 1
            print('%s: %d/%d translated' % (lang, done, len(words)),
                  ('| missing: ' + str(missing)) if missing else '| complete')


if __name__ == '__main__':
    main()
